import tqdm
import os
import openpyxl
import pandas as pd
import numpy as np
import wave
import contextlib
import datetime 
import datetime
import sys


sys.path.insert(0, "/home/aac_s3_test/noisePort_server/04_queries")
from queries import *
from ast import literal_eval
from pathlib import Path
from queries import *
from config import *
from utils_queries import *

import warnings
warnings.filterwarnings("ignore")

def strip_tz(ts):
    if isinstance(ts, pd.Timestamp):
        if ts.tzinfo is not None:
            return ts.tz_localize(None)
        return ts
    return pd.NaT 
    
def read_first_row_excel(path):
    try:
        try:
            df_first_row = pd.read_excel(path, sheet_name='Time History',skiprows=1,nrows=1)
        except:
            df_first_row = pd.read_excel(path, sheet_name='Measurement History')
    
        return df_first_row
    except Exception as e:
        return 0

def get_length_excel(path,sheet_name):
    wb = openpyxl.load_workbook(path,read_only=True,data_only=True)
    ws = wb[f'{sheet_name}'] if sheet_name else wb.active

    last_row_idx = ws.max_row
    
    return last_row_idx

def handle_not_finished_minute(dt):
    """
    Ajusta cualquier pandas.Timestamp o datetime.datetime al inicio de la hora.
    Lo hace truncando: 16:59:59 -> 16:00:00
    #TODO -> gestión de los slops también para sonómetros
    """

    # Si es Timestamp de pandas
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    # Ahora dt es datetime.datetime
    return dt.replace(minute=0, second=0, microsecond=0)

def get_row_indices_by_column(path, sheet_name, column_name, row_content_list, header_row=1):
    
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    excel_length = ws.max_row

    results = np.arange(start = 1,stop=excel_length ,step = 60)
    return results


def get_days_in_df(result_df):

    days_list = []

    for day in result_df['Timestamp']:

        if pd.Timestamp(day).day not in days_list:
            days_list.append(pd.Timestamp(day).day)
    
    return days_list

def process_sonometer_csv(db,csv_path,logger,point,output_folder,processed_txt,count):

        # ------------------------------------
        # 1-    Check if file is in processed txt list
        # ------------------------------------
    if not os.path.exists(processed_txt): open(processed_txt, 'w').close()
    with open(processed_txt) as myfile:
        file = os.path.basename(csv_path)
        if file in myfile.read():
            return
    
    
    try:
     
        # ------------------------------------
        # 2-    Placing columns into starter DFs
        # ------------------------------------

        base_cols = ['LAeq', 'LCeq', 'LAmax', 'LAmin']  
        tail_cols = ['Timestamp', 'Filename', 'Unixtimestamp', 'sensor_id']

        df_measurement = pd.DataFrame()
        df_time = pd.DataFrame()
    except Exception as e:
        logger.error(f'[CSV Sonometer]Error placing columns into starter DFs: {e}')

    try:
        # ------------------------------------
        # 3-    Managing different Svantek output files
        # ------------------------------------
        for csv_file in os.listdir(csv_path):
            
            csv_file_path = os.path.join(csv_path,csv_file)
            fname = Path(csv_file).name  
            
            if 'Time History' in fname:
                df_time = pd.read_csv(csv_file_path,header = 0)
                csv_time_history_path = csv_file_path
                df_final = pd.DataFrame(columns = base_cols + THIRD_OCTAVES + tail_cols)
            elif 'Measurement History' in fname:
                df_measurement = pd.read_csv(csv_file_path,header = 0)
                csv_measurement_path = csv_file_path
                df_final = pd.DataFrame(columns = base_cols + THIRD_OCTAVES + tail_cols)
            elif 'Summary' in fname:
                df_summary = pd.read_csv(csv_file_path,header=0)
                csv_summary_path = csv_file_path
    
    except Exception as e:
        logger.error(f'[CSV Sonometer] Error checking for available sheets in CSV: {e}')

    try:

        
        # ------------------------------------
        # 4-    Reading hourly rows in CSV data sheet
        # ------------------------------------
        
        if not df_measurement.empty:
            
            row_indexs = np.arange(start=1,stop=df_measurement.shape[0],step = 60)
            csv_length = df_measurement.shape[0]
            
            df_measurement = pd.read_csv(csv_measurement_path,header = 0,skiprows = lambda x: x not in np.append(row_indexs,[0]))
            
            

            try:

                df_final[THIRD_OCTAVES] = df_measurement[THIRD_OCTAVES]
                df_final[THIRD_OCTAVES] = df_final[THIRD_OCTAVES].astype(str).replace(',','.',regex=True)
                df_final = df_final.iloc[1:]
                df_final[THIRD_OCTAVES].apply(pd.to_numeric)
            
            except KeyError as e:
                
                df_final[THIRD_OCTAVES_TIME_HISTORY] = df_measurement[THIRD_OCTAVES_TIME_HISTORY]
                df_final[THIRD_OCTAVES_TIME_HISTORY] = df_final[THIRD_OCTAVES_TIME_HISTORY].astype(str).replace(',','.',regex=True)
                df_final = df_final.iloc[1:]
                df_final[THIRD_OCTAVES_TIME_HISTORY].apply(pd.to_numeric)
            
            df_final['LAeq'] = df_measurement['LAeq']
            df_final['LCeq'] = df_measurement[' LCeq']
            
            df_final = df_final.dropna(subset=['LAeq'])
            df_final['LAeq'] = df_final['LAeq'].astype(str).str.replace(',','.')
            
            df_final['LAeq'].apply(pd.to_numeric)
            df_final['LAmax'] = df_final['LAeq'].max()
            df_final['LAmin'] = df_final['LAeq'].min()
            df_final['LAmax'].fillna(df_final['LAmax'],inplace=True)
            df_final['LAmin'].fillna(df_final['LAmin'],inplace=True)
            
            first_row = pd.read_csv(csv_time_history_path,header = None,skiprows=1,nrows=1)
            last_row = df_time.tail(1)
        
            initial_date = handle_not_finished_minute(pd.to_datetime(f'{first_row.iloc[0,2]}' +' '+ f'{first_row.iloc[0,3]}'))
            final_date = pd.to_datetime(f'{last_row.iloc[0,2]}' +' '+ f'{last_row.iloc[0,3]}')

        else:
            
            row_indexs = np.arange(start=1,stop=df_time.shape[0],step = 3600)
            csv_length = df_time.shape[0]
            
            df_time = pd.read_csv(csv_time_history_path,header = 0,skiprows = lambda x: x not in np.append(row_indexs,[0]))
            
            try:

                df_final[THIRD_OCTAVES] = df_time[THIRD_OCTAVES]
                df_final[THIRD_OCTAVES] = df_final[THIRD_OCTAVES].astype(str).replace(',','.',regex=True)
                df_final = df_final.iloc[1:]
                df_final[THIRD_OCTAVES].apply(pd.to_numeric)
            
            except KeyError as e:
                
                df_final[THIRD_OCTAVES_TIME_HISTORY] = df_time[THIRD_OCTAVES_TIME_HISTORY]
                df_final[THIRD_OCTAVES_TIME_HISTORY] = df_final[THIRD_OCTAVES_TIME_HISTORY].astype(str).replace(',','.',regex=True)
                df_final = df_final.iloc[1:]
                df_final[THIRD_OCTAVES_TIME_HISTORY].apply(pd.to_numeric)
        
            df_final['LAeq'] = df_time['LAeq']


            df_final = df_final.dropna(subset=['LAeq'])
            df_final['LAeq'] = df_final['LAeq'].astype(str).str.replace(',','.')
            
            df_final['LAeq'].apply(pd.to_numeric)
            df_final['LAmax'] = df_final['LAeq'].max()
            df_final['LAmin'] = df_final['LAeq'].min()
            df_final['LAmax'].fillna(df_final['LAmax'],inplace=True)
            df_final['LAmin'].fillna(df_final['LAmin'],inplace=True)
            
            first_row = pd.read_csv(csv_time_history_path,header = None,skiprows=1,nrows=1)
            last_row = df_time.tail(1)
        
            initial_date = handle_not_finished_minute(pd.to_datetime(f'{first_row.iloc[0,2]}' +' '+ f'{first_row.iloc[0,3]}'))
            final_date = pd.to_datetime(f'{last_row.iloc[0,2]}' +' '+ f'{last_row.iloc[0,3]}')
    
    except Exception as e:
        logger.error(f'[CSV Sonometer] Error reading hourly rows in CSV data sheet: {e}')

    try:
               
        # ------------------------------------
        # 5-    Asignning desired columns from CSV Summary sheet, and getting timestamp and unixtimestamp info
        # ------------------------------------
        
        df_final['sensor_id'] = df_summary.iloc[2,1]
        df_final['Filename'] = df_summary.iloc[1,1]

        row_content_list = pd.Series(pd.date_range(initial_date,final_date,freq = 'H')
                                            .strftime('%Y-%m-%d %H:%M:%S'))

        df_final['Timestamp'] = row_content_list
        
        ts = pd.to_datetime(df_final['Timestamp'], errors='coerce')  # datetime64[ns], NaT si no se puede parsear
        unixt = (ts.view('int64') // 10**9).astype('Int64')
        unixt[ts.isna()] = pd.NA
        df_final['Unixtimestamp'] = unixt
        
        df_final.sort_values(by='Timestamp',inplace=True)
    except Exception as e:
        logger.error(f'[CSV Sonometer] Error asignning desired columns from CSV Summary sheet: {e}')




    try:

        # ------------------------------------
        # 6-    Saving  whole DF
        # ------------------------------------

        if count == 0:
            file_name = output_folder + f"/{point}_Processed.csv"
            df_final.to_csv(file_name,index = False)
        else:
            file_name = output_folder + f"/{point}_Processed_{count}.csv"
            df_final.to_csv(file_name,index = False)
    
    except Exception as e:
        logger.error(f'[CSV Sonometer]Error eaving per day CSV files from whole DF, and whole DF: {e}')

    try:

        # ------------------------------------
        # 7-    DB upload of result CSV
        # ------------------------------------

        logger.info(f"[CSV Processing] Loading data into TABLE")

        load_data_db(db,file_name,logger,table_name=SONOMETER_TABLE_NAME)
        
        cur = db.cursor()
        cur.execute(f"SELECT COUNT(*) FROM {SONOMETER_TABLE_NAME}")
        n = cur.fetchone()[0]
        logger.info(f"[Acoustics] → {ACOUSTIC_TABLE_NAME} contains {n} rows after LOAD DATA")
        
        cur.close()  
    
    except Exception as e:
        logger.error(f"[CSV Processing] Error at: DB upload of result CSV:{e}")




    try:

        # ------------------------------------
        # 8-    Query and convert results to JSON
        # ------------------------------------
        """
        logger.info("[CSV Processing] Query and Convert Results to JSON")
        avg_results = power_laeq_avg(db,logger,table_name = SONOMETER_TABLE_NAME)
        logger.info(avg_results)

        if avg_results is not None:
            logger.info("[CSV Processing] Power LAeq Average Results:")
            #send_mqtt_data(avg_results, logger)
        else:
            logger.warning("[CSV Processing] No reuslts returned from power_laeq_avg query")
        """
    except Exception as e:
            logger.warning("[CSV Processing] No results returned from power_laeq_avg query.")


    try:
                   
        # ------------------------------------
        # 9-    Update processed folder
        # ------------------------------------
        
        logger.info("")
        update_processed_folder(processed_txt,file)
        processed_txt = load_processed_folder(processed_txt)
        logger.info("[CSV Processing] Added to processed files: %s")

    except Exception as e:
        logger.error(f"[CSV Processing] Error updating processed files: {e}")



def process_sonometer_xlsx(db,xlsx_path,logger,point, output_folder,count,processed_folder_txt):
    
        # ------------------------------------
        # 1-    Check if file is in processed txt list
        # ------------------------------------
    
    
    if not os.path.exists(processed_folder_txt): open(processed_folder_txt, 'w').close()
    with open(processed_folder_txt) as myfile:
        file = os.path.basename(xlsx_path)
        if file in myfile.read():
            return
    try:

        df_first_row = pd.read_excel(xlsx_path, sheet_name='Measurement History',header= None,skiprows=1,nrows=1)
          
        try:

            # ------------------------------------
            # 2-    Getting first and last row info
            # ------------------------------------
                        
            logger.info("[XSLX Processing] Getting info from first and last rows of XLSX file")

            initial_date = handle_not_finished_minute(pd.to_datetime(df_first_row.iloc[0,1],format='%Y-%M-%D'))            
            excel_length = get_length_excel(xlsx_path,'Measurement History')
            
            df_last_row = pd.read_excel(xlsx_path, sheet_name='Measurement History',header= None,skiprows=excel_length-1,nrows=1)
            final_date = pd.to_datetime(df_last_row.iloc[0,1],format='%Y-%M-%D')
            
            df_last_row = df_last_row.drop(np.r_[0:43,79:151],axis=1)
            df_last_row.dropna(axis=1,inplace=True)
            df_last_row.columns = THIRD_OCTAVES
        
            df_first_row = df_first_row.drop(np.r_[0:43,79:151],axis=1)
            df_first_row.dropna(axis=1,inplace=True)
            df_first_row.columns = THIRD_OCTAVES
        
        except Exception as e:
            logger.error(f"[XSLX Processing] Error reading first or last rows of excel file: {xlsx_path}")


        try:

            # ------------------------------------
            # 3-    Getting content and index info from the desired hourly range
            # ------------------------------------           
            
            logger.info("[XSLX Processing] Getting content and index info from the desired hourly range")

            row_content_list = pd.Series(pd.date_range(initial_date,final_date,freq = 'H')
                                            .strftime('%Y-%m-%d %H:%M:%S'))
            
            row_indexs = get_row_indices_by_column(xlsx_path,sheet_name='Measurement History',column_name = 'Time',row_content_list = row_content_list)

        except Exception as e:
            logger.error(f"[XSLX Processing] Error getting index and context info from hourly range from file: {xlsx_path}")


        try:

            # ------------------------------------
            # 4-    Reading excel file and naming columns
            # ------------------------------------
                        
            logger.info("[XSLX Processing] Reading excel file and naming columns")
            df_result = pd.read_excel(xlsx_path,sheet_name = 'Measurement History',usecols = "AR:CA",header = 0,skiprows = lambda x: x not in row_indexs )
            df_summary = pd.read_excel(xlsx_path,sheet_name = 'Summary')

            df_result.columns = THIRD_OCTAVES
            df_result[['LA','LC']] = pd.read_excel(xlsx_path,sheet_name = 'Measurement History',usecols = "AH:AI",header = 0,skiprows = lambda x: x not in row_indexs )
            df_result['sensor_id'] = df_summary.iloc[2,1]
            df_result['LAmax'] = df_result['LA'].max()
            df_result['LAmin'] = df_result['LC'].min()
            df_result['LAmax'].fillna(df_result['LAmax'], inplace=True)
            df_result['LAmin'].fillna(df_result['LAmin'], inplace=True)        
            df_result['sensor_id'].fillna(df_result['sensor_id'], inplace=True) 
        
        except Exception as e:
            logger.error(f"[XSLX Processing] Error reading excel file and naming columns from file: {xlsx_path}")

        try:

            # ------------------------------------
            # 5-    Implementing data into columns from previously retrieved data and reordering columns
            # ------------------------------------        
            
            logger.info("[XSLX Processing] Implementing data into columns from previously retrieved data")
            
            df_last_row['Timestamp'] = final_date
            
            df_result = pd.concat([df_result,df_last_row],ignore_index = True)

            df_result['Timestamp'] = row_content_list
            df_result['Timestamp'] = pd.to_datetime(df_result['Timestamp'], errors='coerce')
            df_result['Filename'] = os.path.basename(xlsx_path)
            df_result['Filename'].fillna(os.path.basename(xlsx_path)) 
            df_result['Unixtimestamp'] = (df_result['Timestamp'].view('int64') // 10**9).astype('Int64')

            columns = ['sensor_id','Filename','Timestamp','Unixtimestamp','LA','LC','LAmax','LAmin']          
            df_result = df_result[columns + THIRD_OCTAVES]
            
            df_result.sort_values(by='Timestamp')            

        except Exception as e:
            logger.error(f"[XSLX Processing] Error implementing data into columns of xlsx file: {xlsx_path}")

        try:

            # ------------------------------------
            # 6-    Saving per day CSV files from whole DF, and whole DF
            # ------------------------------------

            logger.info("[XSLX Processing] Saving per day CSV files from whole DF, and whole DF")

            days = get_days_in_df(df_result)
            output_folder_days = output_folder +'/daily' f'/{point}'
            output_folder_whole = output_folder + f"/{point}_Processed_{count}.csv"
            os.makedirs(output_folder_whole, exist_ok=True)
            os.makedirs(output_folder_days, exist_ok=True)
            
            for day in days:
                
                day_df = df_result.loc[df_result['Timestamp'].dt.day == int(day)].copy()
                filename = os.path.join(output_folder_days, f"day{day}_{point}_Processed_{count}.csv") 
                day_df.to_csv(filename, index=False)

            df_result.to_csv(output_folder_whole)

        except Exception as e:
            logger.error(f"[XSLX Processing] Error splitting whole DF into per day DFs:{e}")


        try:

            # ------------------------------------
            # 7-    DB upload of result CSV
            # ------------------------------------

            logger.info(f"[XSLX Processing] Loading data into TABLE")
            
            load_data_db(db,output_folder_whole,logger,table_name = SONOMETER_TABLE_NAME)
            
            cur = db.cursor()
            cur.execute(f"SELECT COUNT(*) FROM {SONOMETER_TABLE_NAME}")
            n = cur.fetchone()[0]
            logger.info(f"[Acoustics] → {ACOUSTIC_TABLE_NAME} contains {n} rows after LOAD DATA")
            cur.close()                
        
        except Exception as e:
            logger.error(f"[XSLX Processing] Error at: DB upload of result CSV:{e}")


        try:
         
            # ------------------------------------
            # 8-    Query and convert results to JSON
            # ------------------------------------
            
            logger.info("[XLSX Processing] Query and Convert Results to JSON")
            avg_results = power_laeq_avg(db,logger,table_name = SONOMETER_TABLE_NAME)
            logger.info(avg_results)
            
            if avg_results is not None:
                logger.info("[XLSX Processing] Power LAeq Average Results:")
                send_mqtt_data(avg_results, logger)
            else:
                logger.warning("[XLSX Processing] No reuslts returned from power_laeq_avg query")
        
        except Exception as e:
            logger.warning("[XLSX Processing] No results returned from power_laeq_avg query.")

        try:
                        
            # ------------------------------------
            # 8- Update processed folder
            # ------------------------------------
            
            logger.info("")
            update_processed_folder(processed_folder_txt,day)
            load_processed_folder(processed_folder_txt)
            logger.info("[XLSX Processing] Added to processed files: %s",day)

        except Exception as e:
            logger.error(f"[XLSX Processing] Error updating processed files: {e}")
    except Exception as e:
        logger.error(f"Error reading Measurement History sheet from {xlsx_path} ,  trying Time History sheet")


def process_acoustic_folder(db,logger,folder_days,all_info,query_folder,processed_folder,processed_folder_txt):
    if not os.path.exists(processed_folder_txt): open(processed_folder_txt, 'w').close()
    for day in tqdm.tqdm(folder_days, desc="[Acoustics] Processing days", unit="day"):
        
        with open(processed_folder_txt) as myfile:
            if day in myfile.read():
                return
      
        try:

            # ------------------------------------
            # 1-Taking day string to save the concat file
            # ------------------------------------
            
            day_str = day.split("/")[-1]
            logger.info("[Acoustics] Processing day_hour: %s", day_str)
            logger.info("[Acoustics] Processing: %s", day)
        
        except Exception as e:
            logger.error(f"[Acoustics] Error processing day: {e}")
            continue


        try:
            
            # ------------------------------------
            # 2-Appending to list csv files in csv per day folder
            # ------------------------------------
                        
            csv_files = os.listdir(day)
            csv_files = [csv_file for csv_file in csv_files if csv_file.endswith(".csv")]
            logger.info("[Acoustics] Processing CSV files in %s: %s", day, csv_files)
            csv_files = [os.path.join(day, csv_file) for csv_file in csv_files]
        
        except Exception as e:
            logger.error(f"[Acoustics] Error listing CSV files: {e}")
            continue


        try:

            # ------------------------------------
            # 3-Concatenation of csv files for one hour processing
            # ------------------------------------
            
            logger.info("")
            logger.info("[Acoustics] Trying to concatenate the csv files to process one hour of audio data recordings")
            df_day = pd.concat([pd.read_csv(csv_file) for csv_file in csv_files], ignore_index=True)
        
        except Exception as e:
            logger.error(f"[Acoustics] Error concatenating CSV files: {e}")
            continue

        try:            
            
            # ------------------------------------
            # 4-ordering by timestamp, selecting columns,truncate first minute if not = 00 and turning the result into a csv so we can use it
            # ------------------------------------

            df_day = df_day.sort_values(by=["Timestamp"])
            df_day["LA"] = df_day["LA"].round(1)

            # -------------------------------
            # Truncar los minutos iniciales incompletos
            # -------------------------------
            if not df_day.empty:
                logger.info("[Acoustics] Converting 'Timestamp' to datetime and handling timezone.")
                
                # 1. Convertir a datetime (esto es lo que puede crear datetimes tz-aware)
                df_day['Timestamp'] = pd.to_datetime(df_day['Timestamp'], errors='coerce')
                
                # Intentamos la naivización usando el accesor .dt de Pandas.
                try:
                    # Si la columna ya es datetime y tiene TZ, la naivizamos.
                    if df_day['Timestamp'].dt.tz is not None:
                        logger.warning("[Acoustics] Timestamp column is tz-aware. Naivizing using tz_localize(None).")
                        # Esta línea es la que necesitamos para quitar la TZ sin cambiar la hora.
                        df_day['Timestamp'] = df_day['Timestamp'].dt.tz_localize(None)
                except AttributeError:
                    # Si falla al acceder a .dt (porque es una Serie 'object' mezclada)
                    # o si el paso anterior no fue suficiente, naivizamos usando una función lambda
                    # que maneja los objetos datetime individuales de Python.
                    logger.warning("[Acoustics] Direct naivization failed. Applying lambda to strip timezone from elements.")

                    df_day['Timestamp'] = df_day['Timestamp'].apply(strip_tz)

                # 3. Forzar el tipo final para asegurar que .dt.second funcione
                # Ya que hemos quitado la TZ, este .astype no debería fallar con ValueError.
                df_day['Timestamp'] = df_day['Timestamp'].astype('datetime64[ns]')
                
                # 4. La lógica de truncado, que ahora es segura
                mask_sec0 = df_day['Timestamp'].dt.second == 0
                
                if mask_sec0.any():
                    first_sec0_idx = mask_sec0.idxmax()
                    df_day = df_day.loc[first_sec0_idx:]
                    logger.info("[Acoustics] Truncated leading incomplete minute.")
                else:
                    df_day = df_day.iloc[0:0]
                    logger.warning("[Acoustics] No records starting at second :00 found. DataFrame truncated to empty.")
                
            
            columns = ["LA", "LC", "LZ", "LAmax", "LAmin"] + THIRD_OCTAVES_SECOND_FORMAT + [
                                "Timestamp","Filename","Unixtimestamp","id_micro"]
            csv_concat_path = os.path.join(query_folder, f"{day_str}.csv")
            
            # 5. Limpiar NaT y formatear antes de guardar
            df_day = df_day.dropna(subset=['Timestamp'])
            df_day['Timestamp'] = df_day['Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')

            df_day = df_day[columns]
            df_day.to_csv(csv_concat_path, index=False)
            logger.info("[Acoustics] Concatenated CSV files, saved as: %s", csv_concat_path)
        
        except Exception as e:
            logger.error(f"[Acoustics] Error saving concatenated CSV file: {e}")
            continue


        try:

            # ------------------------------------
            # 5-Loading ACOUSTIC csv into the DB table
            # ------------------------------------
            
            logger.info("")
            logger.info("[Acoustics] Loading data into TABLE")
            load_data_db(db, csv_concat_path, logger,table_name=ACOUSTIC_TABLE_NAME)
            cur = db.cursor()
            cur.execute(f"SELECT COUNT(*) FROM {ACOUSTIC_TABLE_NAME}")
            n = cur.fetchone()[0]
            logger.info(f"[Acoustics] → {ACOUSTIC_TABLE_NAME} contains {n} rows after LOAD DATA")
            cur.close()
        
        except Exception as e:
            logger.error(f"[Acoustics] Error loading data into database: {e}")
            continue

        """
        try:
            
            # ------------------------------------
            # 6-query and convert results to json
            # ------------------------------------            
            
                        logger.info("")
            logger.info("[Acoustics] Query and Convert Results to JSON")
            avg_results = power_laeq_avg(db, logger)
            logger.info(avg_results)
            for result in avg_results:
                result["day_path"] = day

            logger.info("[Acoustics] Power LAeq Average Results:")
            logger.info(avg_results)

            if avg_results is not None:
                logger.info("[Acoustics] Power LAeq Average Results:")
                send_mqtt_data(avg_results, logger)
            else:
                logger.warning("[Acoustics] No results returned from power_laeq_avg query.")
            
        except Exception as e:
            logger.error(f"[Acoustics] Error querying and converting results to JSON: {e}")
            continue

        all_info.append(avg_results)
        """

        try:          

            # ------------------------------------
            # 7-Update processed folder
            # ------------------------------------

            logger.info("")
            update_processed_folder(processed_folder_txt, day)
            processed_folder = load_processed_folder(processed_folder_txt)
            logger.info("[Acoustics] Added to processed files: %s", day)
        
        except Exception as e:
            logger.error(f"[Acoustics] Error updating processed files: {e}")
            continue



def process_pred_folder(db,logger,folder_days, all_info, query_folder, processed_folder, processed_folder_txt):
    if not os.path.exists(processed_folder_txt): open(processed_folder_txt, 'w').close()
    for day in tqdm.tqdm(folder_days, desc="[Predictions] Processing days", unit="day"):
        with open(processed_folder_txt) as myfile:
            if day in myfile.read():
                continue


        try:

            # ------------------------------------
            # 1-    Taking day string to save the concat file
            # ------------------------------------

            day_str = day.split("/")[-1]
            logger.info("[Predictions] Processing day_hour: %s", day_str)
            logger.info("[Predictions] Processing: %s", day)
        
        except Exception as e:
            logger.error(f"[Predictions] Error processing day: {e}")
            continue

        try:

            # ------------------------------------
            # 2-    Appending to list csv files in csv per day folder
            # ------------------------------------


            csv_files = os.listdir(day)
            csv_files = [csv_file for csv_file in csv_files if csv_file.endswith("1.0.csv")]
            logger.info("[Predictions] CSV files in %s: %s", day, csv_files)
            csv_files = [os.path.join(day, csv_file) for csv_file in csv_files]
        
        except Exception as e:
            logger.error(f"[Predictions] Error listing CSV files: {e}")
            continue

        try:

            # ------------------------------------
            # 3-    Concatenation of csv files for one hour processing
            # ------------------------------------


            logger.info("")
            logger.info("[Predictions] Trying to concatenate the csv files to process one hour of audio data recordings")
            df_day = pd.concat([
                pd.read_csv(
                    csv_file,
                    converters={'class': literal_eval, 'probability': literal_eval}
                )
                for csv_file in csv_files
                if csv_file.endswith("1.0.csv")
            ], ignore_index=True)
        
        except Exception as e:
            logger.error(f"[Predictions] Error concatenating CSV files: {e}")
            continue

        try:

            # ------------------------------------
            # 4-    Ordering by timestamp 
            # exploding prediction and probability columns  
            # turning the result into a csv so we can use it
            # rearranging df columns so it fits in the table
            # ------------------------------------


            df_day = df_day.sort_values(by=["Timestamp"])
            df_day["prediction1"],df_day["prediction2"],df_day["prediction3"] = zip(*list(df_day['class'].values))
            df_day['probability1'],df_day['probability2'],df_day['probability3'] = zip(*list(df_day['probability'].values))


            # -------------------------------
            # Truncar los minutos iniciales incompletos
            # -------------------------------
            if not df_day.empty:
                # Asegurarse de que 'Timestamp' sea datetime
                df_day['Timestamp'] = pd.to_datetime(df_day['Timestamp'], errors='coerce')

                # Crear máscara: True donde el segundo es 0
                mask_sec0 = df_day['Timestamp'].dt.second == 0
                if mask_sec0.any():  # si hay algún segundo = 0
                    first_sec0_idx = mask_sec0.idxmax()  # obtiene el índice de la primera fila con segundo = 0
                    df_day = df_day.loc[first_sec0_idx:]  # truncar hasta esa fila inclusive
                else:
                    # Si no hay ninguna fila con segundo=0, opcionalmente vaciar todo
                    df_day = df_day.iloc[0:0]


            df_out = df_day.rename(columns={
                'prediction1': 'Prediction_1',
                'prediction2': 'Prediction_2',
                'prediction3': 'Prediction_3',
                'probability1': 'Prob_1',
                'probability2': 'Prob_2',
                'probability3': 'Prob_3'
            })

            cols = ['Prediction_1','Prediction_2','Prediction_3',
                    'Prob_1','Prob_2','Prob_3',
                    'Filename','Timestamp']
            df_out = df_out[cols]
            csv_concat_path = os.path.join(query_folder, f"{day_str}.csv")
            logger.info("[Predictions] Concatenated CSV file path: %s", csv_concat_path)
            df_out.to_csv(csv_concat_path, index=False)
            logger.info("[Predictions] Concatenated CSV files, saved as: %s", csv_concat_path)
        
        except Exception as e:
            logger.error(f"[Predictions] Error saving concatenated CSV file: {e}")
            continue

        try:

            # ------------------------------------
            # 5-    Loading PREDICTIONS csv into the DB table
            # ------------------------------------


            logger.info("")
            logger.info("[Predictions] Loading data into TABLE")
            load_data_db(db, csv_concat_path, logger,table_name=PREDICT_TABLE_NAME)
            cur = db.cursor()
            cur.execute(f"USE {DATABASE_NAME}")
            cur.execute(f"SELECT COUNT(*) FROM {PREDICT_TABLE_NAME}")
            n = cur.fetchone()[0]
            logger.info(f"[Predictions] → {PREDICT_TABLE_NAME} contains {n} rows after LOAD DATA")
            cur.close()
        
        except Exception as e:
            logger.error(f"[Predictions] Error loading data into database: {e}")
            continue
        """
        try:

            # ------------------------------------
            # 6-    Query and convert results to json
            # ------------------------------------


            logger.info("")
            logger.info("[Predictions] Query and Convert Results to JSON")
            avg_results = power_laeq_avg(db, logger)
            logger.info(avg_results)
            for result in avg_results:
                result["day_path"] = day

            logger.info("[Predictions] Power LAeq Average Results:")
            logger.info(avg_results)

            if avg_results is not None:
                logger.info("[Predictions] Power LAeq Average Results:")
                send_mqtt_data(avg_results, logger)
            else:
                logger.warning("[Predictions] No results returned from power_laeq_avg query.")
        
        except Exception as e:
            logger.error(f"[Predictions] Error querying and converting results to JSON: {e}")
            continue
        """
        try:

            # ------------------------------------
            # 7-    Update processed folder 
            # ------------------------------------


            logger.info("")
            update_processed_folder(processed_folder_txt, day)
            processed_folder = load_processed_folder(processed_folder_txt)
            logger.info("[Predictions] Added to processed files: %s", day)
        
        except Exception as e:
            logger.error(f"[Predictions] Error updating processed files: {e}")
            continue


def process_wav_folder(db,logger,folder_days, all_info, query_folder, processed_folder, processed_folder_txt):
    if not os.path.exists(processed_folder_txt): open(processed_folder_txt, 'w').close()
    for day in tqdm.tqdm(folder_days, desc="[Wave Files] Processing days", unit="day"):
        
        with open(processed_folder_txt) as myfile:
            if day in myfile.read():
                continue


        try:

            # ------------------------------------
            # 1-    Taking day string to save the concat file
            # ------------------------------------
            
            day_str = day.split("/")[-1]
            logger.info("[Wave Files] Processing day_hour: %s", day_str)
            logger.info("[Wave Files] Processing: %s", day)
        except Exception as e:
            logger.error(f"[Wave Files] Error processing day: {e}")
            continue

        try:

            # ------------------------------------
            # 2-    Reading wav time lengths from wav folder
            # ------------------------------------


            duration = []
            
            for wavfile in os.listdir(day):
                if wavfile.endswith(".wav"):
                    with contextlib.closing(wave.open(os.path.join(day,wavfile),'r')) as f:
                            frames = f.getnframes()
                            rate = f.getframerate()
                            duration_wav = frames / float(rate)
                            duration.append(duration_wav)
        except Exception as e:
            logger.error(f"[Wave Files] Error listing CSV files: {e}")
            continue

        try:

            # ------------------------------------
            # 3-    Creating csvs with filename, timestamp and duration
            # ------------------------------------
            
            logger.info("")
            logger.info("[Wave Files] Trying to create csv files with filename, timestamp and duration")
            df_day = pd.DataFrame(columns=['Filename','Timestamp','Duration'])

            df_day['Filename'] = os.listdir(day)
            df_day['Duration'] = duration
            df_day['Timestamp'] = pd.to_datetime(
                df_day['Filename'].astype(str).str.replace('.wav', '', regex=False),
                format='%Y%m%d_%H%M%S',
                errors='raise'  
            )

        except Exception as e:
            logger.error(f"[Predictions] Error concatenating CSV files: {e}")
            continue

        try:
            
            # ------------------------------------
            # 4-    Ordering by timestamp 
            # saving csv to wav_files_query folder
            # ------------------------------------

            df_day = df_day.sort_values(by=["Timestamp"])
            csv_concat_path = os.path.join(query_folder, f"{day_str}.csv")
            df_day.to_csv(os.path.join(query_folder, f"{day_str}.csv"), index=False)
            logger.info(f"[Wave Files] Concatenated CSV files, saved as:{csv_concat_path}" )

        except Exception as e:
            logger.error(f"[Wave Files] Error saving concatenated CSV file: {e}")
            continue

        try:

            # ------------------------------------
            # 5-    Loading PREDICTIONS csv into the DB table
            # ------------------------------------

            logger.info("")
            logger.info("[Wave Files] Loading data into TABLE")
            
            load_data_db(db, csv_concat_path, logger,table_name=WAV_TABLE_NAME)
            
            cur = db.cursor(buffered=True)
            cur.execute(f"USE {DATABASE_NAME}")
            cur.execute(f"SELECT COUNT(*) FROM {WAV_TABLE_NAME}")
            n = cur.fetchone()[0]
            logger.info(f"[Wave Files] → {WAV_TABLE_NAME} contains {n} rows after LOAD DATA")
            cur.close()
        
        except Exception as e:
            logger.error(f"[Wave Files] Error loading data into database: {e}")
            continue
        """
        try:

            # ------------------------------------
            # 6-    Query and convert results to json
            # ------------------------------------

            logger.info("")
            logger.info("[Wave Files] Query and Convert Results to JSON")
            avg_results = power_laeq_avg(db, logger)
            logger.info(avg_results)
            
            for result in avg_results:
                result["day_path"] = day

            logger.info("[Wave Files] Power LAeq Average Results:")
            logger.info(avg_results)

            if avg_results is not None:
                
                logger.info("[Wave Files] Power LAeq Average Results:")
                send_mqtt_data(avg_results, logger)
            
            else:
                
                logger.warning("[Wave Files] No results returned from power_laeq_avg query.")
        
        except Exception as e:
            logger.error(f"[Wave Files] Error querying and converting results to JSON: {e}")
            continue
        """
        # ------------------------------------
        # 7-    Update processed folder 
        # ------------------------------------
        try:
            logger.info("")
            update_processed_folder(processed_folder_txt, day)
            processed_folder = load_processed_folder(processed_folder_txt)
            logger.info("[Wave Files] Added to processed files: %s", day)
        except Exception as e:
            logger.error(f"[Wave Files] Error updating processed files: {e}")
            continue



def process_sonometer_folder(db,logger,files_folder,query_folder,processed_sonometers_txt):
        
        output_folder = query_folder.replace('sonometer_files','sonometer_acoustics_query')
        
        for point in tqdm.tqdm(os.listdir(files_folder), desc="[Sonometers] Processing Points"):
                
            point_folder = os.path.join(files_folder,point)
            file_count = 0
            
            lxt_files = [f for f in os.listdir(point_folder) if f.endswith('.xlsx')]                    
            csvs_files = [f for f in os.listdir(point_folder) if f.endswith('_CSV')]
            
            if csvs_files != []:
                
                for file in tqdm.tqdm(csvs_files, desc=f"Processing files in {point}"):
                    
                    file_path = os.path.join(point_folder,file)                        
                    logger.info(f"[SONOMETER] -> Processing file: {file_path}")                            
                    process_sonometer_csv(db,file_path,logger,point,output_folder,processed_sonometers_txt,file_count)                                                    
                    file_count += 1
                    logger.info(f"[SONOMETER] -> Processed data saved at: {output_folder}")
            
            else:

                for file in tqdm.tqdm(lxt_files,desc = f'Processing files in {point}'):

                    file_path = os.path.join(files_folder,point, file)
                    logger.info(f"[SONOMETER] -> Processing file: {file_path}")                            
                    process_sonometer_xlsx(db,file_path,logger,point,output_folder,file_count,processed_sonometers_txt)
                    file_count += 1
                    logger.info(f"[SONOMETER] -> Processed data saved at: {output_folder}")